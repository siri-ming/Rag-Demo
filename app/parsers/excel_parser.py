"""Excel spreadsheet parser using openpyxl."""

from pathlib import Path

import openpyxl

from .base import BaseParser, Document


class ExcelParser(BaseParser):
    """
    Parser for Excel (.xlsx/.xls) files.

    Converts each worksheet into a text representation with headers and rows.
    """

    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def _sheet_to_text(self, sheet) -> str:
        """Convert a single worksheet to text format."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return ""

        lines = []
        # First row as header
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        lines.append(" | ".join(headers))
        lines.append("-" * 40)

        # Data rows
        for row in rows[1:]:
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                lines.append(" | ".join(cells))

        return "\n".join(lines)

    def parse(self, file_path: str | Path) -> list[Document]:
        file_path = Path(file_path)
        documents = []

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        file_name = file_path.name

        try:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text = self._sheet_to_text(sheet)

                if text.strip():
                    documents.append(
                        Document(
                            text=text,
                            metadata={
                                "source": file_name,
                                "file_type": "excel",
                                "sheet_name": sheet_name,
                            },
                        )
                    )
        finally:
            wb.close()

        return documents
